import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

#####################################
# ASSUMPTIONS SHEET
#####################################
ws_assumptions = wb.create_sheet('Assumptions')

# Key Inputs
assumptions_data = [
    ['WAL-MART DCF VALUATION ANALYSIS', '', ''],
    ['February 2010', '', ''],
    ['', '', ''],
    ['KEY ASSUMPTIONS', '', ''],
    ['', '', ''],
    ['Market Data (as of Feb 2010)', '', ''],
    ['Current Stock Price', 53.48, '$'],
    ['Shares Outstanding (millions)', 3866, ''],
    ['Market Capitalization (millions)', 206731.28, '$'],
    ['', '', ''],
    ['Risk-Free Rate (10-year Treasury)', 3.68, '%'],
    ['Market Risk Premium', 5.05, '%'],
    ['Beta (Bloomberg)', 0.66, ''],
    ['Cost of Equity (CAPM)', 7.01, '%'],
    ['', '', ''],
    ['Historical Performance', '', ''],
    ['FY2010 Revenue (millions)', 405046, '$'],
    ['FY2010 EBIT (millions)', 22066, '$'],
    ['FY2010 Tax Rate', 32.35, '%'],
    ['FY2010 NOPAT (millions)', 14926, '$'],
    ['FY2010 Depreciation (millions)', 7157, '$'],
    ['FY2010 CapEx (millions)', 12184, '$'],
    ['FY2010 Change in NWC (millions)', 4368, '$'],
    ['FY2010 Free Cash Flow (millions)', 5531, '$'],
    ['', '', ''],
    ['STAGE 1: High Growth Period (Years 1-5)', '', ''],
    ['Revenue Growth Rate', 6.0, '%'],
    ['EBIT Margin', 5.45, '%'],
    ['Tax Rate', 32.35, '%'],
    ['Depreciation as % of Revenue', 1.77, '%'],
    ['CapEx as % of Revenue', 3.01, '%'],
    ['NWC as % of Revenue Change', 25.0, '%'],
    ['', '', ''],
    ['STAGE 2: Stable Growth Period (Perpetuity)', '', ''],
    ['Perpetual Growth Rate', 3.0, '%'],
    ['Terminal EBIT Margin', 5.5, '%'],
    ['Terminal Depreciation as % of Revenue', 2.0, '%'],
    ['Terminal CapEx as % of Revenue', 2.5, '%'],
    ['Terminal NWC as % of Revenue Change', 20.0, '%'],
]

for i, row in enumerate(assumptions_data, 1):
    for j, val in enumerate(row, 1):
        cell = ws_assumptions.cell(row=i, column=j, value=val)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif i in [4, 6, 16, 26, 33]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

ws_assumptions.column_dimensions['A'].width = 40
ws_assumptions.column_dimensions['B'].width = 15
ws_assumptions.column_dimensions['C'].width = 10

#####################################
# DCF MODEL SHEET
#####################################
ws_dcf = wb.create_sheet('DCF Model')

# Parameters
revenue_2010 = 405046
growth_stage1 = 0.06
growth_terminal = 0.03
ebit_margin_stage1 = 0.0545
ebit_margin_terminal = 0.055
tax_rate = 0.3235
depr_pct_stage1 = 0.0177
capex_pct_stage1 = 0.0301
nwc_pct_stage1 = 0.25
depr_pct_terminal = 0.02
capex_pct_terminal = 0.025
nwc_pct_terminal = 0.20
cost_of_equity = 0.0701

# Build projections
years = list(range(0, 7))
revenues = [revenue_2010]
for i in range(1, 7):
    if i <= 5:
        revenues.append(revenues[-1] * (1 + growth_stage1))
    else:
        revenues.append(revenues[-1] * (1 + growth_terminal))

ebit = []
for i, rev in enumerate(revenues):
    if i <= 5:
        ebit.append(rev * ebit_margin_stage1)
    else:
        ebit.append(rev * ebit_margin_terminal)

nopat = [e * (1 - tax_rate) for e in ebit]

depreciation = []
for i, rev in enumerate(revenues):
    if i <= 5:
        depreciation.append(rev * depr_pct_stage1)
    else:
        depreciation.append(rev * depr_pct_terminal)

capex = []
for i, rev in enumerate(revenues):
    if i <= 5:
        capex.append(rev * capex_pct_stage1)
    else:
        capex.append(rev * capex_pct_terminal)

delta_nwc = [0]
for i in range(1, 7):
    delta_rev = revenues[i] - revenues[i-1]
    if i <= 5:
        delta_nwc.append(delta_rev * nwc_pct_stage1)
    else:
        delta_nwc.append(delta_rev * nwc_pct_terminal)

fcf = []
for i in range(7):
    fcf.append(nopat[i] + depreciation[i] - capex[i] - delta_nwc[i])

# Discount factors
discount_factors = [(1 / (1 + cost_of_equity) ** i) for i in range(7)]

# PV of cash flows
pv_fcf = [fcf[i] * discount_factors[i] for i in range(7)]

# Terminal value (at end of year 5)
terminal_fcf = fcf[6]
terminal_value = terminal_fcf / (cost_of_equity - growth_terminal)
pv_terminal_value = terminal_value * discount_factors[5]

# Enterprise value and equity value
sum_pv_fcf = sum(pv_fcf[1:6])
enterprise_value = sum_pv_fcf + pv_terminal_value
equity_value = enterprise_value
value_per_share = equity_value / 3866

# Write to sheet
dcf_headers = ['Item', 'FY2010', 'FY2011', 'FY2012', 'FY2013', 'FY2014', 'FY2015', 'FY2016']
ws_dcf.append(dcf_headers)

data_rows = [
    ['Revenue', revenues[0], revenues[1], revenues[2], revenues[3], revenues[4], revenues[5], revenues[6]],
    ['Revenue Growth %', '', growth_stage1*100, growth_stage1*100, growth_stage1*100, growth_stage1*100, growth_stage1*100, growth_terminal*100],
    ['', '', '', '', '', '', '', ''],
    ['EBIT', ebit[0], ebit[1], ebit[2], ebit[3], ebit[4], ebit[5], ebit[6]],
    ['EBIT Margin %', ebit[0]/revenues[0]*100, ebit_margin_stage1*100, ebit_margin_stage1*100, ebit_margin_stage1*100, ebit_margin_stage1*100, ebit_margin_stage1*100, ebit_margin_terminal*100],
    ['Tax @ 32.35%', ebit[0]*tax_rate, ebit[1]*tax_rate, ebit[2]*tax_rate, ebit[3]*tax_rate, ebit[4]*tax_rate, ebit[5]*tax_rate, ebit[6]*tax_rate],
    ['NOPAT', nopat[0], nopat[1], nopat[2], nopat[3], nopat[4], nopat[5], nopat[6]],
    ['', '', '', '', '', '', '', ''],
    ['Add: Depreciation', depreciation[0], depreciation[1], depreciation[2], depreciation[3], depreciation[4], depreciation[5], depreciation[6]],
    ['Less: CapEx', capex[0], capex[1], capex[2], capex[3], capex[4], capex[5], capex[6]],
    ['Less: Increase in NWC', delta_nwc[0], delta_nwc[1], delta_nwc[2], delta_nwc[3], delta_nwc[4], delta_nwc[5], delta_nwc[6]],
    ['', '', '', '', '', '', '', ''],
    ['Free Cash Flow', fcf[0], fcf[1], fcf[2], fcf[3], fcf[4], fcf[5], fcf[6]],
    ['', '', '', '', '', '', '', ''],
    ['Discount Factor', discount_factors[0], discount_factors[1], discount_factors[2], discount_factors[3], discount_factors[4], discount_factors[5], discount_factors[6]],
    ['PV of FCF', pv_fcf[0], pv_fcf[1], pv_fcf[2], pv_fcf[3], pv_fcf[4], pv_fcf[5], pv_fcf[6]],
]

for row in data_rows:
    ws_dcf.append(row)

ws_dcf.append(['', '', '', '', '', '', ''])
ws_dcf.append(['VALUATION SUMMARY', '', '', '', '', '', ''])
ws_dcf.append(['', '', '', '', '', '', ''])
ws_dcf.append(['Sum of PV of FCF (Years 1-5)', sum_pv_fcf, '', '', '', '', ''])
ws_dcf.append(['Terminal FCF (Year 6)', terminal_fcf, '', '', '', '', ''])
ws_dcf.append(['Terminal Value', terminal_value, '', '', '', '', ''])
ws_dcf.append(['PV of Terminal Value', pv_terminal_value, '', '', '', '', ''])
ws_dcf.append(['', '', '', '', '', '', ''])
ws_dcf.append(['Enterprise Value', enterprise_value, '', '', '', '', ''])
ws_dcf.append(['Less: Net Debt', 0, '', '', '', '', ''])
ws_dcf.append(['Equity Value', equity_value, '', '', '', '', ''])
ws_dcf.append(['', '', '', '', '', '', ''])
ws_dcf.append(['Shares Outstanding (millions)', 3866, '', '', '', '', ''])
ws_dcf.append(['Value per Share', value_per_share, '', '', '', '', ''])
ws_dcf.append(['Current Market Price', 53.48, '', '', '', '', ''])
ws_dcf.append(['Upside/(Downside)', value_per_share - 53.48, '', '', '', '', ''])
ws_dcf.append(['Upside/(Downside) %', (value_per_share - 53.48) / 53.48 * 100, '', '', '', '', ''])

# Format
for row in ws_dcf.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

ws_dcf.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dcf.column_dimensions[col].width = 15

#####################################
# SENSITIVITY ANALYSIS SHEET
#####################################
ws_sens = wb.create_sheet('Sensitivity Analysis')

ws_sens.append(['SENSITIVITY ANALYSIS: VALUE PER SHARE', '', '', '', '', '', ''])
ws_sens.append(['', '', '', '', '', '', ''])
ws_sens.append(['Perpetual Growth Rate vs. Cost of Equity', '', '', '', '', '', ''])
ws_sens.append(['', '', '', '', '', '', ''])

# Create sensitivity table
growth_rates = [0.02, 0.025, 0.03, 0.035, 0.04]
costs_of_equity = [0.06, 0.065, 0.0701, 0.075, 0.08]

# Header row
header_row = ['Cost of Equity / Growth Rate'] + [f'{g*100:.1f}%' for g in growth_rates]
ws_sens.append(header_row)

# Calculate sensitivity values
for coe in costs_of_equity:
    row_data = [f'{coe*100:.2f}%']
    for g in growth_rates:
        if g < coe:
            # Recalculate with new assumptions
            tv_sens = terminal_fcf / (coe - g)
            pv_tv_sens = tv_sens / ((1 + coe) ** 5)
            ev_sens = sum_pv_fcf + pv_tv_sens
            value_sens = ev_sens / 3866
            row_data.append(round(value_sens, 2))
        else:
            row_data.append('N/A')
    ws_sens.append(row_data)

ws_sens.column_dimensions['A'].width = 30
for col in ['B', 'C', 'D', 'E', 'F', 'G']:
    ws_sens.column_dimensions[col].width = 15

# Save
wb.save('walmart_dcf_analysis.xlsx')
print(f'Excel file created successfully')
print(f'Calculated value per share: ${value_per_share:.2f}')
print(f'Current market price: $53.48')
print(f'Upside/(Downside): ${value_per_share - 53.48:.2f} ({(value_per_share - 53.48) / 53.48 * 100:.1f}%)')
